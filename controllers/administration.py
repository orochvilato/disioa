# -*- coding: utf-8 -*-
# essayez quelque chose comme
import json
def getPrivate():
    name = request.args(0)
    import os
    f = open(os.path.join(request.folder,'private','data',name))
    response.headers['ContentType'] ="application/octet-stream";
    response.headers['Content-Disposition']="attachment; filename="+name
    
    return response.stream(f,chunk_size=4096)

    
def updateElectionTour():
    mdb.deputes.update_one({'depute_shortid':'anneblanc'},{'$set':{'depute_election.adversaires':[{'nom':"M. André AT (retiré)",'voix':0}]}})
    mdb.deputes.update_many({'depute_election.adversaires.0': { "$exists": False } }, {'$set':{'depute_election.tour':1}})
    mdb.deputes.update_many({'depute_election.adversaires.0': { "$exists": True } }, {'$set':{'depute_election.tour':2}})
def test():
    from openpyxl import load_workbook
    from cStringIO import StringIO
    import requests
    
    r =requests.get("https://docs.google.com/spreadsheets/d/1lZ5aMIaglRh6_BK66AYkXzi9guJaMij9RlpSoO4dnIw/export?format=xlsx&id=1lZ5aMIaglRh6_BK66AYkXzi9guJaMij9RlpSoO4dnIw")
    f = StringIO(r.content)
    wb = load_workbook(f)
    ws = wb['scrutins']
    elts = []
    for j,row in enumerate(ws.iter_rows(min_row=2)):
        elts.append([a.value for a in row])
    scrutins_cles = {}
    for e in elts:
        if e[0] != None:
            scrutins_cles[int(e[0])] = dict(num=e[0],nom=e[1],desc=e[2],theme=e[3],lien=e[4])
    return BEAUTIFY(scrutins_cles)
    
def error():
    1/0
def updateLogs():
    from geoip import geolite2
    import pycountry
    from user_agents import parse
    for l in mdb.logs.find({'$or':[{'geoip':None},{'agent_pretty':None}]}):
        match = geolite2.lookup(l['client'])
        pays = ""
        if match:
            country = match.country
            if country:
                pays = pycountry.countries.get(alpha_2=country).name
        agent_pretty = str(parse(l['agent']))
        mdb.logs.update({'_id':l['_id']},{'$set':{'geoip':pays,'agent_pretty':agent_pretty}})

def logs():
    updateLogs()
    return dict(logs=mdb.logs.find().sort('date',-1).limit(200))
