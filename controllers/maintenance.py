# -*- coding: utf-8 -*-
# essayez quelque chose comme

import os
import re
import json
import datetime
import csv
import requests
from tools import normalize
from pymongo import UpdateOne
from bson.son import SON
from tools import normalize
import xmltodict
client = pymongo.MongoClient('mongodb://writer:discordinsoumisrw@localhost:27017/obsass')
mdb = client.obsass
output_path = os.path.join(request.folder, 'private', 'scrapy')

def corrections_manuelles():
    mdb.scrutins.update_one({'scrutin_id':'15_134'},{'$set':{'scrutin_groupe':'FI'}})
    mdb.scrutins.update_one({'scrutin_id':'15_135'},{'$set':{'scrutin_groupe':'LR'}})
    mdb.votes.update_many({'scrutin_id':'15_134'},{'$set':{'scrutin_groupe':'FI'}})
    mdb.votes.update_many({'scrutin_id':'15_135'},{'$set':{'scrutin_groupe':'LR'}})
    
def groupeFromNom(nom):
    def gfromn():
        gfn  = {}
        for d in mdb.deputes.find({},{'depute_id':1,'depute_nom':1,'groupe_abrev':1}):
            g = d['groupe_abrev']
            nom =d['depute_nom'].split(' ')
            gfn[normalize(nom[0])+normalize(''.join(nom[2:]))] = {'g':g,'id':d['depute_id']}
        return gfn
    gfn = cache.ram('gfn',lambda:gfromn(),time_expire=3600)
    return gfn[nom]

def update_deputesphotos():
    from base64 import b64encode
    import requests
    for d in mdb.deputes.find():
        #path = os.path.join(request.folder, 'static/images/deputes/%s.jpg' % source['depute_id'])
        #content = open(path).read()
        print d['depute_shortid']
        content =requests.get('http://www2.assemblee-nationale.fr/static/tribun/15/photos/'+d['depute_uid'][2:]+'.jpg').content
        photo = b64encode(content)
        mdb.deputes.update_one({'depute_uid':d['depute_uid']},{'$set':{'depute_photo':photo}})
        
def get_amendements():
    result = cache.disk('amendements',lambda:fetch_amendements(),time_expire=0)
    return result

def fetch_amendements():
    
    leg = request.args(0) or '15'
    from HTMLParser import HTMLParser
    html = HTMLParser()
    gfn  = {}
    for d in mdb.deputes.find({},{'depute_id':1,'depute_nom':1,'groupe_abrev':1}):
        g = d['groupe_abrev']
        nom =d['depute_nom'].split(' ')
        gfn[d['depute_id']]= {'g':g,'id':d['depute_id'],'nom':d['depute_nom']}
        gfn[normalize(nom[0])+normalize(''.join(nom[2:]))] = {'g':g,'id':d['depute_id'],'nom':d['depute_nom']}
        gfn[normalize(nom[0])+normalize(''.join(nom[-1]))] = {'g':g,'id':d['depute_id'],'nom':d['depute_nom']}
    
    amendements = []
    nb = 1000
    count = 1000
    start = 1
    stats = {}
    stats_deps = {}
    notfound = []
    while (count==nb):
        r = requests.get('http://www2.assemblee-nationale.fr/recherche/query_amendements?typeDocument=amendement&leg=%s&idExamen=&idDossierLegislatif=&missionVisee=&numAmend=&idAuteur=&premierSignataire=false&idArticle=&idAlinea=&sort=&dateDebut=&dateFin=&periodeParlementaire=&texteRecherche=&rows=%d&format=json&tri=ordreTexteasc&start=%d&typeRes=liste' % (leg,nb,start))
        amds = r.json()
        fields = amds['infoGenerales']['description_schema'].split('|')
        count = len(amds[u'data_table'])
        start += nb
        for a in amds[u'data_table']:
            amd = dict((fields[i],_a) for i,_a in enumerate(a.split('|')))
            amd['signataires_ids'] = []
            amd['signataires_groupes'] = []
            for i,s in enumerate(amd['signataires'].replace(' et ', ',').split(',')):
                s = normalize(html.unescape(s))
                if s in gfn.keys():
                    if not gfn[s]['nom'] in stats_deps:
                        stats_deps[gfn[s]['nom']] = {'signes':0,'rediges':0,'adoptes':0,'groupe':gfn[s]['g']}
                    if i==0:
                        stats_deps[gfn[s]['nom']]['rediges'] += 1
                        if amd['sort']==u'Adopté':
                            stats_deps[gfn[s]['nom']]['adoptes'] += 1
                    stats_deps[gfn[s]['nom']]['signes'] += 1
                    amd['signataires_ids'].append(gfn[s]['id'])
                    amd['signataires_groupes'].append(gfn[s]['g'])
                else:
                    notfound.append(s)
            if amd['signataires_groupes']:
                g = amd['signataires_groupes'][0]
            else:
                g = 'Gouvernement'
            if not g in stats:
                stats[g] = dict()
            stats[g][amd['sort']] = stats[g].get(amd['sort'],0)+1
            amendements.append(amd)
    totaux = {}
    for d in stats_deps.keys():
        mdb.deputes.update({'depute_nom':d},{'$set':{'depute_amendements':stats_deps[d]}})
        if stats_deps[d]['rediges'] > 0:
            stats_deps[d]['pct'] = round(100*float(stats_deps[d]['adoptes'])/stats_deps[d]['rediges'],2)
    for g in stats.keys():
        stats[g]['total'] = sum(stats[g].values())
    for f in ['total',u'Adopt\xe9',u'Rejet\xe9',u'Retir\xe9',u'Tomb\xe9',u'Non renseign\xe9',u'Non soutenu']:
        totaux[f] = sum([stats[g].get(f,0) for g in stats.keys()])
    return BEAUTIFY(notfound)
    return dict(stats=stats,totaux=totaux,stats_deps=sorted(stats_deps.iteritems(),key=lambda x:(x[1]['signes'],x[1]['signes']),reverse=True))
    

def correctData():
    mdb.scrutins.update({'scrutin_id':'15_135'},{'$set':{'scrutin_groupe':'LR'}})
    mdb.scrutins.update({'scrutin_id':'15_134'},{'$set':{'scrutin_groupe':'FI'}})

def updateDeputesStatsElection():
    from openpyxl import load_workbook
    path = os.path.join(request.folder, 'private','data','legislatives.xlsx')
    wb = cache.ram('circo.xls',lambda:load_workbook(path),time_expire=0)
    circos = dict((d['depute_departement_id'],d['depute_departement']) for d in mdb.deputes.find())
    deputes = {}
    for d in mdb.deputes.find():
        deputes[d['depute_circo_id']] = deputes.get(d['depute_circo_id'],[])+[d['depute_shortid']]
    
    ws = wb['Circo. leg. T2']
    updates = {}
    cv = { "ZA":"971", "ZB":"972", "ZC":"973", "ZD":"974", "ZS":"975", "ZM":"976", "ZX":"977", "ZW":"986", "ZP":"987", "ZN":"988", "ZZ":"999"}
    
    for r in list(ws.iter_rows())[3:]:
        advs = []
        dep = ('000'+cv.get(str(r[0].value),str(r[0].value)))[-3:]
        t = dep +"-"+('00'+str(r[2].value))[-2:]
       
        i=28
        while i<len(r) and r[i].value:
            advs.append({'nom':{'M':'M. ','F':'Mme '}[r[i].value]+r[i+2].value+' '+r[i+1].value+' ('+r[i+3].value+')','voix':r[i+4].value})
            i += 9

        for d in deputes[t]:
            #mdb.deputes.update({'depute_shortid':d},{'$set':{'depute_election':{'inscrits':r[4].value, 'adversaires':advs, 'abstentions':r[5].value,'votants':r[7].value,'blancs':r[9].value,'nuls':r[12].value,'exprimes':r[15].value,'voix':r[23].value,'depute_election.nuance':r[22].value}}})
            mdb.deputes.update({'depute_shortid':d},{'$set':{'depute_election.nuance':r[22].value}})
    return json.dumps(advs)

def weekly_job():
    updateAssemblee()
    

def daily_job1():
    updateScrutins()
    updateSessions()
def daily_job2():
    updateCommissions()
    get_amendements()
def daily_job3():
    updateScrutinsCles()
    updateAllStats()
def daily_job():
    daily_job1()
    daily_job2()
    daily_job3()
    
# ---------------------
# Helpers
# ---------------------
def updatePhotos():
    upd = []
    for d in mdb.deputes.find():
        r =requests.get('http://www2.assemblee-nationale.fr/static/tribun/15/photos/'+d['depute_uid'][2:]+'.jpg')
        with open(os.path.join(request.folder, 'static', 'images','deputes',d['depute_id']+'.jpg'),'w') as f:
            f.write(r.content)
        upd.append(d['depute_id'])
    return BEAUTIFY(upd)

def formatpct(n,d,prec=1):
    return round(float(n)/d,prec)
def updateCircos():
    paris = ['077-01','077-02','077-03','077-04','077-05','077-06','077-07','077-08','077-09','077-10','077-11','078-01','078-02','078-03','078-04','078-05','078-06','078-07','078-08','078-09','078-10','078-11','078-12','091-01','091-02','091-03','091-04','091-05','091-06','091-07','091-08','091-09','091-10','095-01','095-02','095-03','095-04','095-05','095-06','095-07','095-08','095-09','095-10','092-01','092-02','092-03','092-04','092-05','092-06','092-07','092-08','092-09','092-10','092-11','092-12','092-13','093-01','093-02','093-03','093-04','093-05','093-06','093-07','093-08','093-09','093-10','093-11','093-12','094-01','094-02','094-03','094-04','094-05','094-06','094-07','094-08','094-09','094-10','094-11','075-01','075-02','075-03','075-04','075-05','075-06','075-07','075-08','075-09','075-10','075-11','075-12','075-13','075-14','075-15','075-16','075-17','075-18']
    villes = ['075-01','075-02','075-03','075-04','075-05','075-06','075-07','075-08','075-09','075-10','075-11','075-12','075-13','075-14','075-15','075-16','075-17','075-18','069-01','069-02','069-03','069-04','069-06','069-07','069-12','069-14','013-01','013-02','013-03','013-04','013-05','013-06','013-07','044-01','044-02','033-01','033-02','033-03','031-01','031-04','006-01','006-03','067-01','067-02','067-03','059-01','059-02','059-04','059-07','059-08','059-09','059-10']
    circopath = os.path.join(request.folder, 'private','circonscriptions_france2.svg')
    svg = xmltodict.parse(open(circopath,'r'))
    for c in svg['svg']['path']:
        circo = dict(ville=(c['@id'] in villes),paris=(c['@id'] in paris),carte='france',d=c['@d'],id=c['@id'],dep=c['@id'].split('-')[0],title=c['title']['#text'],desc=c['desc']['#text'])
        mdb.circonscriptions.update({'id':c['@id']},{'$set':circo},upsert=True)
    return "ok"


def updateHemicycle():
    hemipath = os.path.join(request.folder, 'views/svg/hemicyclelight.svg')
    svg = xmltodict.parse(open(hemipath,'r'))
    places_deputes = dict((d['depute_place'],d) for d in mdb.deputes.find())
    for p in svg['svg']['a']:
        place = p['path']['@id'][1:]
        depute = places_deputes.get(place,None)
        if depute:
            p['@href'] = "{{ =base_url[1:]+'/"+depute['depute_shortid']+"' if 'base_url' in locals() else '#' }}"
            p['@title'] = "place %s : %s (%s)" % (depute['depute_place'],depute['depute_nom'],depute['groupe_abrev'])
            p['path']['@class'] = "place %s" % depute['groupe_abrev']
            p['title']={'#text':p['@title']}
            if '@target' in p.keys():
                del p['@target']
    
    import codecs
    file = codecs.open(hemipath, "w", "utf-8")
    file.write(xmltodict.unparse(svg,pretty=True))
        
    return 'ok'

def launchScript(name,params=""):
    fp = os.path.join(request.folder, 'private/scripts', name +'.py '+output_path+' '+params)

    did_scrape = True
    if not 'debug' in request.args:
        did_scrape = True if os.system(fp) else False

    return did_scrape

def getData(name,id):
    fp = os.path.join(request.folder, 'private/data/', name +'.csv')
    with open(fp) as csvfile:
        reader = csv.DictReader(csvfile,delimiter='|')
        result = dict((row[id].decode('utf8'),dict((k,v.decode('utf8')) for k,v in row.iteritems())) for row in reader)
    return result
def addData(name,elts):
    fp = os.path.join(request.folder, 'private/data/', name +'.csv')
    with open(fp,'a') as f:
        f.write('|'.join([e.encode('utf8') for e in elts])+'\n')

def getJson(name):
    return json.loads(open(output_path+'/'+name+'.json','r').read())



def genNuages(_mots):
    mts = {}
    for lex in _mots.keys():
        if _mots[lex]:
            mots = [ [mot,count] for mot,count in sorted(_mots[lex].items(),key=lambda x:x[1],reverse=True) if not mot in nuages_excl][:200]
            if mots:
                mx = mots[0][1]
                mn = mots[-1][1]-1
                coef = 12000/sum([len(mot)*(float(count-mn)/(mx-mn))**2 for mot,count in mots])
                mts[lex] = [ [mot,int(coef*float(count-mn)/(mx-mn))] for mot,count in mots]

    return mts

# ---------------------
# Imports de données
# ---------------------

def buildIndexes():
    mdb.interventions.ensure_index([('itv_date',pymongo.DESCENDING),('session_id',pymongo.ASCENDING),('itv_n',pymongo.ASCENDING)])
    mdb.interventions.reindex()
    mdb.votes.ensure_index([('scrutin_num',pymongo.DESCENDING)])
    mdb.votes.reindex()
    return 
    # Deputes
    mdb.deputes.drop_indexes()
    mdb.deputes.ensure_index([("depute_nom", pymongo.TEXT)],default_language='french')
    mdb.deputes.ensure_index([("depute_nom_tri", pymongo.ASCENDING)])
    #mdb.deputes.ensure_index([("depute_nom_sort", pymongo.DESCENDING)])
    #mdb.deputes.reindex()
    #mdb.scrutins.ensure_index([("scrutin_fulldesc",pymongo.TEXT)],default_language='french')
    #mdb.scrutins.reindex()
    #mdb.votes.ensure_index([("scrutin_fulldesc",pymongo.TEXT)],default_language='french')
    #mdb.votes.reindex()
    #mdb.interventions.ensure_index([("itv_contenu_texte",pymongo.TEXT)],default_language='french')
    #mdb.interventions.reindex()
    sortfields = ['stats.positions.exprimes','stats.positions.dissidence','stats.nbitvs','stats.nbmots','stats.compat.FI','stats.compat.REM',
                  'stats.amendements.rediges','stats.amendements.adoptes','stats.commissions.present','stats.election.exprimes','stats.election.inscrits']
    ranks = ['exprimes','dissidence','compatFI','compatREM','nbitvs','nbmots','nbamendements','pctamendements','pctcommissions','pctinscrits','pctexprimes']
    for sf in sortfields:
        mdb.deputes.ensure_index([(sf,pymongo.ASCENDING)])
    for o in 'up','down':
        for r in ranks:
            mdb.deputes.ensure_index([('stats.nonclasse',pymongo.ASCENDING),('stats.ranks.%s.%s' % (o,r),pymongo.ASCENDING)])
    
    mdb.deputes.reindex()
    
def updateCommissions():
    import datetime
    dhc=mdb.deputes.find({'depute_commissions_historique':{'$ne':[]}},{'depute_id':1,'depute_commissions_historique':1,'depute_mandat_debut':1,'depute_mandat_fin':1})
    
    histcom = {}
    historique = []
    current = {}
    for d in dhc:
        fin =  datetime.datetime.strptime(d['depute_mandat_fin'] or '01/01/2999','%d/%m/%Y')
        debut =  datetime.datetime.strptime(d['depute_mandat_debut'],'%d/%m/%Y')
        cfin = datetime.datetime.strptime(d['depute_commissions_historique'][0][1],'%d/%m/%Y') + datetime.timedelta(days=1)
        if cfin<debut:
            cfin = debut
        current[d['depute_id']] = cfin.strftime('%d/%m/%Y')
        for cdebut,cfin,comm in d['depute_commissions_historique']:
            _cdebut = datetime.datetime.strptime(cdebut,'%d/%m/%Y')
            _cfin = datetime.datetime.strptime(cfin,'%d/%m/%Y')
            if (_cdebut>=debut and _cdebut<=fin):
                if comm[0:3]=='OMC':
                    mdb.presences.delete_many({'$and':[{'depute_id':d['depute_id']},{'presence_etat':'absent'},{'presence_date':{'$lte':_cfin,'$gte':_cdebut}},{'commission_id':{'$ne':comm}}]})
                
                historique.append([d['depute_id'],cdebut, cfin, comm])
    
    for h in historique:
        if not h[3] in histcom.keys():
            histcom[h[3]] = {}
        if not h[0] in histcom[h[3]].keys():
            histcom[h[3]][h[0]]=[]
        histcom[h[3]][h[0]].append([h[1],h[2]])
        
    open(output_path+'/'+'comparams.json','w').write(json.dumps(dict(current=current,histo=histcom)))
    launchScript('commissions','comparams.json')
    
    presences  =getJson('presences')
    #amendements  =getJson('amendements')
    commissions = getJson('commissions')
    groupe_depute = dict((d['depute_id'],d['groupe_abrev']) for d in mdb.deputes.find({},{'depute_id':1,'groupe_abrev':1}))
    for com in commissions:
        mdb.commissions.update_one({'commission_id':com['commission_id']},{'$set':com},upsert=True)
    
    import datetime
    for pres in presences:
        pres['depute_id'] = pres['depute_id'].replace('\r','').replace('8','').replace('6','')
                    
        pres['groupe_abrev'] = groupe_depute.get(pres['depute_id'],groupe_depute.get('m.'+pres['depute_id'],groupe_depute.get('mme'+pres['depute_id'],'Nope')))
        if pres['groupe_abrev']=='Nope':
            1/0
        
        pres['presence_date'] = datetime.datetime.strptime(pres['presence_date'],'%Y-%m-%d %H:%M')
        mdb.presences.update_one({'presence_id':pres['presence_id']},{'$set':pres},upsert=True)
        
    # presences / depute
    pgroup = {}
    pgroup['n'] = {'$sum':1}
    pgroup['_id'] = { 'depute_id':'$depute_id', 'commission_id':'$commission_id','etat':'$presence_etat'}
    pipeline = [{'$group':pgroup}]
    presences_deputes = {}
    for p in mdb.presences.aggregate(pipeline):
        depid = p['_id']['depute_id']
        commissionid = p['_id']['commission_id'].split('_')[1]
        etat = p['_id']['etat']
        if not depid in presences_deputes.keys():
            presences_deputes[depid]={}
        if not commissionid in presences_deputes[depid].keys():
            presences_deputes[depid][commissionid] = {'present':0,'absent':0,'excuse':0,'total':0}
        presences_deputes[depid][commissionid][etat] += p['n']
        presences_deputes[depid][commissionid]['total'] += p['n']
    for d in mdb.deputes.find({},{'depute_commissions':1,'depute_id':1}):
        if d['depute_id'] in presences_deputes.keys():
            d['depute_presences_commissions'] = presences_deputes[d['depute_id']]
            mdb.deputes.update_one({'depute_id':d['depute_id']},{'$set':d})
    
    # presences / par groupe
    pgroup = {}
    pgroup['n'] = {'$sum':1}
    pgroup['_id'] = { 'groupe':'$groupe_abrev', 'etat':'$presence_etat'}
    pipeline = [{'$group':pgroup}]
    presences_groupes = {}
    for p in mdb.presences.aggregate(pipeline):
        groupeabrev = p['_id']['groupe']
        etat = p['_id']['etat']
        if not groupeabrev in presences_groupes.keys():
            presences_groupes[groupeabrev]={'present':0,'absent':0,'excuse':0}
        presences_groupes[groupeabrev][etat] += p['n']
    #for amd in amendements:
    #    amd['groupe_abrev'] = groupe_depute.get(amd['auteur_id'],amd['auteur_id'])
    #    mdb.amendements.update_one({'amendement_id':amd['amendement_id']},{'$set':amd},upsert=True)
    #for depid,pres in presences_deputes.iteritems():
    #    mdb.deputes.presence
    
    return "ok"
def updateAssemblee():
    updateHATVP()
    #updateDeputyWatch()
    launchScript('assemblee')
    deputes = getJson('deputes')
    deputywatch = getJson('deputywatch')
    departements = getData('departements','departement')
    professions = getData('deputes_professions','id')
    professions2 = getData('deputes_professions','profession')
    hatvp = getJson('hatvp')
    csp_incomplet = []
    for d in deputes:
        d['depute_sexe'] = 'Homme' if d['depute_nom'][0:2]=='M.' else 'Femme'
        d['depute_id'] = normalize(d['depute_nom'])
        d['depute_shortid'] = d['depute_id'][2:] if '.' in d['depute_id'] else d['depute_id'][3:]
        d['depute_age'] = int((datetime.datetime.now() - datetime.datetime.strptime(d['depute_ddn'],'%d/%m/%Y')).days / 365.25)
        d['depute_classeage'] = '%d-%d ans' % ((d['depute_age']/10)*10,(1+(d['depute_age']/10))*10)
        d['depute_deputywatch'] = deputywatch.get(d['depute_id'],None)
        d['depute_hatvp'] = hatvp.get(d['depute_id'],[])
        d['depute_region'] = departements[d['depute_departement']]['region']
        d['depute_typeregion'] = departements[d['depute_departement']]['typeregion']
        d['depute_departement_id'] = departements_ids[d['depute_departement']]
        d['depute_circo_id'] = d['depute_departement_id']+'-'+('00'+d['depute_circo'])[-2:]
        m = professions.get(d['depute_uid'],None)
        if d['depute_hatvp']:
            d['depute_nomcomplet'] = "%s %s" % (d['depute_hatvp'][0]['nom'],d['depute_hatvp'][0]['prenom'])
        if not m:
            m = professions2.get(d['depute_profession'],None)
        csp = ""
        if m:
            csp = m['csp']

        if not csp and not d['depute_uid'] in professions.keys():
            addData('deputes_professions',(d['depute_uid'],d['depute_nom'],d['depute_profession'],''))

        d['depute_csp'] = csp
        mdb.deputes.update_one({'depute_uid': d['depute_uid']}, {'$set': d}, upsert = True)

    groupes = getJson('groupes')
    for g in groupes:
        membres = mdb.deputes.find({'groupe_uid':g['groupe_uid']})
        g['groupe_membres'] = [ dict(qualite=m['groupe_qualite'],uid=m['depute_uid'],actif = m['depute_actif']) for m in membres]
        g['groupe_nbmembres'] = len([ m for m in g['groupe_membres'] if m['actif']])
        mdb.groupes.update_one({'groupe_uid':g['groupe_uid']},{'$set':g}, upsert= True)

    # initialise les champs stats
    mdb.deputes.update_many({'stats':None},{'$set':{'stats':{'nbmots':0,'nbitvs':0}}})
    mdb.groupes.update_many({'stats':None},{'$set':{'stats':{'nbmots':0,'nbitvs':0}}})
    mdb.groupes.update_many({'groupe_mots':None},{'$set':{'groupe_mots':{}}})
    mdb.deputes.update_many({'depute_mots':None},{'$set':{'depute_mots':{}}})

    return mdb.deputes.find().count()

def updateDeputyWatch():
    launchScript('deputywatch')

def updateHATVP():
    types_doc = {
     'dia': u'Déclaration d’intérêts et d’activités',
     'diam': u'Déclaration de modification substantielle des intérêts et des activités',
     'di': u'Déclaration d’intérêts',
     'dim': u'Déclaration de modification substantielle des intérêts',
     'dsp': u'Déclaration de situation patrimoniale',
     'dspm': u'Déclaration de modification substantielle de situation patrimoniale',
     'dspfm': u'Déclaration de modification substantielle de situation patrimoniale',
     'appreciation': u'Appréciation de la HATVP'
    }

    import requests
    import csv
    import json
    from cStringIO import StringIO
    from tools import normalize
    r = requests.get('http://www.hatvp.fr/files/open-data/liste.csv')
    #f = StringIO(r.content[3:])
    f = StringIO(r.content)
    csv = csv.DictReader(f, delimiter=';', quotechar='"')
    declarations = {}
    for row in csv:
        
        drow = dict((k,v.decode('utf8') if isinstance(v,basestring) else v) for k,v in row.iteritems())
        
        id = normalize(drow[u'civilite']+' '+drow['prenom']+' '+drow['nom'])
        
        
        code = drow['url_dossier'].split('/')[-1]
        drow['docurl'] = 'http://www.hatvp.fr/fiche-nominative/?declarant='+code
        declarations[id] = declarations.get(id,[])+ [drow]

    with open(output_path+'/hatvp.json','w') as f:
        f.write(json.dumps(declarations))

def updateSessions():
    if request.args(0)=='rebuild':
        exclude = []
    else:
        exclude = mdb.interventions.distinct('session_id')
    launchScript('sessions',"'%s'" % (json.dumps(exclude)))
    
    sessions = getJson('sessions')
    scrutins = list(mdb.scrutins.find())
    for s in scrutins:
        if s['scrutin_typedetail']=='amendement' and 'scrutin_ref' in s.keys():
            cptrd = s['scrutin_ref']['urlCompteRenduRef'].split('#')[0]
            if cptrd in sessions.keys():
                balises = sessions[cptrd]
                bal = balises.get(s['scrutin_ref']['numAmend'],'')
                s['scrutin_ref']['urlCompteRenduRef'] = cptrd + '#' + bal
                mdb.scrutins.update_one({'scrutin_id': s['scrutin_id']}, {'$set': {'scrutin_ref': s['scrutin_ref']}})

    deputes = dict((d['depute_id'],{'uid':d['depute_uid'],'shortid':d['depute_shortid']}) for d in mdb.deputes.find({},{"depute_id":1,"depute_uid":1,"depute_shortid":1}))
    deputes_id = dict((v['uid'],k) for k,v in deputes.iteritems())
    interventions = getJson('interventions')
    ops = []
    for itv in interventions:
        for n,depid in enumerate(itv['depute_id'].split(u'|')):
            new_itv=dict(itv)
            new_itv['depute_id'] = depid
            nid = None
            if depid in deputes.keys():
                nid = deputes[depid]['uid']
            if nid and nid != new_itv['depute_uid']:
                new_itv['depute_uid'] = nid
                new_itv['itv_tribun'] = True
            elif not nid and deputes_id.get(new_itv['depute_uid'],None):
                new_itv['depute_id'] = deputes_id[new_itv['depute_uid']]
            if new_itv['itv_president']==new_itv['depute_id']:
                new_itv['itv_president'] = True
            else:
                new_itv['itv_president'] = False
            new_itv['itv_id'] = "%s%d" % (new_itv['itv_id'],n)
            if new_itv['depute_id'] in deputes.keys():
                new_itv['depute_shortid'] = deputes[new_itv['depute_id']]['shortid']
            ops.append(UpdateOne({'itv_id':new_itv['itv_id']},{'$set':new_itv}, upsert=True))
    if ops:
        mdb.interventions.bulk_write(ops)

    if request.args(0)!='rebuild':
        updateInterventionsNuages()


def updateInterventionsNuages():
    lexiques_path = os.path.join(request.folder, 'private','lexiques.json')

    with open(lexiques_path,'r') as f:
        lexiques = json.loads(f.read())

    lexiquenoms = lexiques['NOM']
    #lexiquenoms.update(lexiques['ADV'])
    lexiquenoms.update(lexiques['ADJ'])
    lexiqueverbes = lexiques['VER']
    _lexiques = {'noms':lexiquenoms,'verbes':lexiqueverbes}
    _lex = {'noms':set(_lexiques['noms'].keys()),
            'verbes':set(_lexiques['verbes'].keys()) }

    def countWords(txt):
        words = {'noms':{},'verbes':{}}
        txt = txt.replace('\n',' ').replace('!','').replace('?','').replace('.',' ').replace(':',' ').replace(',',' ').replace(';',' ').replace('-',' ').replace('  ',' ').replace(u'\xa0','').replace(u'\u2019',' ').lower().split(' ')
        for lex in _lexiques.keys():
            _words = [ _lexiques[lex][x] for x in txt if x in _lex[lex] ]
            for w in _words:
                if not w.strip():
                    continue
                if not w in words[lex].keys():
                    words[lex][w] = 1
                else:
                    words[lex][w] += 1

        return words

    def addWords(w1,w2):
        for k in w2.keys():
            if not k in w1.keys():
                w1[k] = w2[k]
            else:
                for mot,n in w2[k].iteritems():
                    w1[k][mot] = w1[k].get(mot,0) + n
    
    groupe_depute = dict((d['depute_id'],(d['groupe_abrev'],d['depute_nom'])) for d in mdb.deputes.find({},{'depute_id':1,'groupe_abrev':1,'depute_nom':1}))
    deputes = dict((a['depute_id'],{ 'groupe_abrev':a['groupe_abrev'],'stats.nbmots':0, 'stats.nbitvs':0, 'depute_mots':{}, 'depute_nuages':{}}) for a in mdb.deputes.find())
    groupes = dict((g['groupe_abrev'], {'groupe_mots':{},'groupe_nuages':{},'stats.nbmots':0,'stats.nbitvs':0}) for g in mdb.groupes.find())
    interventions = mdb.interventions.find()
    for itv in interventions:
        updt ={'mots':countWords(itv['itv_contenu_texte'])}
        if itv['depute_id'] in deputes.keys():
            _dep = deputes[itv['depute_id']]
            _dep['stats.nbmots'] += itv['itv_nbmots']
            _dep['stats.nbitvs'] += 1
            addWords(_dep['depute_mots'],updt['mots'])
            updt.update({'groupe_abrev':groupe_depute[itv['depute_id']][0],'depute_nom':groupe_depute[itv['depute_id']][1]})
            _grp = groupes[_dep['groupe_abrev']]
            _grp['stats.nbmots'] += itv['itv_nbmots']
            _grp['stats.nbitvs'] += 1
            addWords(_grp['groupe_mots'],updt['mots'])

        mdb.interventions.update({'itv_id':itv['itv_id']},{'$set':updt})

    for gid,g in groupes.iteritems():
        g['groupe_nuages'] = genNuages(groupes[gid]['groupe_mots'])
        mdb.groupes.update_one({'groupe_abrev':gid},
                               {'$set':g})

    for uid,d in deputes.iteritems():
        d['depute_nuages'] = genNuages(deputes[uid]['depute_mots'])
        mdb.deputes.update_one({'depute_id':uid},
                                   {'$set':d})
    return "ok"


def updateScrutins():
    _scrutins = {}
    if 'rebuild' in request.args:
        mdb.votes.remove()
        mdb.scrutins.remove()

    # TODO : fonctionnalité import PDF à conserver ?
    #_scrutins = get_scrutinsPDFs()
    #fp = os.path.join(current.request.folder, 'private/scripts', 'scrutins.py')
    #did_scrape = True if os.system(fp) else False

    scrutins_complets = [ s['scrutin_id'] for s in list(mdb.scrutins.find({'$and':[{'scrutin_dossier':{'$ne':'N/A'}},{'scrutin_type':{'$ne':'N/A'}}]}))]
    launchScript('scrutins',"'"+json.dumps(scrutins_complets)+"'")
    _scrutins.update(getJson('scrutins'))

    votes = []
    scrutins = []

    deputes = dict((d['depute_id'],d) for d in mdb.deputes.find())

    def act_votedata(id,position,posori=None):
        d = deputes[id]
        return {'depute_uid':d['depute_uid'],
                'depute_shortid':d['depute_shortid'],
                'depute_nom':d['depute_nom'],
                'groupe_uid': d['groupe_uid'],
                'groupe_abrev': d['groupe_abrev'],
                'depute_age':d['depute_age'],
                'depute_classeage':d['depute_classeage'],
                'depute_region':d['depute_region'],
                'depute_typeregion':d['depute_typeregion'],
                'depute_departement':d['depute_departement'],
                'depute_csp':d['depute_csp'],
                'depute_sexe':d['depute_sexe'],
                'vote_position':position,
                'vote_position_ori':posori }

    for s in _scrutins.values():
        desc = s['desc'].replace('. [','.')
        types = s.get('libelleType','N/A')
        fulldesc = u"Scrutin n° %d du %s : %s" % (s[u'num'],s[u'date'],desc)
        scrutin = { 'scrutin_num':s['num'],
                    'scrutin_id':s['id'],
                    'scrutin_desc':desc,
                    'scrutin_fulldesc':fulldesc,
                    'scrutin_date':s['date'],
                    'scrutin_type':s.get('type','N/A'),
                    'scrutin_typeLibelle':types,
                    'scrutin_dossier':s.get('idDossier','N/A'),
                    'scrutin_dossierLibelle':s.get('libelleDossier','N/A'),
                    'scrutin_typedetail':s['typedetail'],
                    'scrutin_ok':s['ok']
                   }

        exprimes = []
        for p in s['votes'].keys():
            for v in s['votes'][p]:
                vote = dict(scrutin)
                depute = deputes.get('m.'+v,deputes.get('mme'+v,deputes.get(v,'PB')))
                if depute == 'PB':
                    print v.encode('utf8')
                else:
                    exprimes.append(depute['depute_id'])
                    if p=='nonVotant':
                        continue
                    corr = s['corrections'].get('m.'+v,s['corrections'].get('mme'+v,False))
                    if corr:
                        position = corr
                        posori = p
                    else:
                        position = p
                        posori = p
                    vote.update(act_votedata(depute['depute_id'],position,posori))
                    vote['vote_id'] = "%d_%s" % (s['num'],depute['depute_id'])
                    votes.append(vote)
        for depcorr,pos in s['corrections'].iteritems():
            if not depcorr in exprimes:
                vote = dict(scrutin)
                vote.update(act_votedata(depcorr,pos,'absent'))
                vote['vote_id'] = "%d_%s" % (s['num'],depcorr)
                votes.append(vote)
                exprimes.append(depcorr)

        for id in list(set(deputes.keys())-set(exprimes)):
            datescrutin = datetime.datetime.strptime(s['date'],'%d/%m/%Y')
            depute = deputes[id]

            if (depute['depute_mandat_debut'] and datetime.datetime.strptime(depute['depute_mandat_debut'],'%d/%m/%Y')>datescrutin) or (depute['depute_mandat_fin'] and datetime.datetime.strptime(depute['depute_mandat_fin'],'%d/%m/%Y')<datescrutin):

                continue

            vote = dict(scrutin)
            vote.update(act_votedata(id,'absent'))
            vote['vote_id'] = "%d_%s" % (s['num'],id)
            votes.append(vote)
        if 'reference' in s.keys():
            scrutin.update({'scrutin_ref':s['reference']})
        scrutin.update({  'scrutin_lienscrutin':s['scrutinlien'],
                            'scrutin_liendossier':s['dossierlien'] })
        scrutins.append(scrutin)



    scrutins_incomplets = [ s['scrutin_id'] for s in list(mdb.scrutins.find({'$or':[{'scrutin_dossier':'N/A'},{'scrutin_type':'N/A'}]}))]

    mdb.votes.remove({'scrutin_id':{'$in':scrutins_incomplets}})
    mdb.scrutins.remove({'scrutin_id':{'$in':scrutins_incomplets}})

    # stockage des votes et des scrutins dans la base
    if votes:
        mdb.votes.create_index([('vote_id', pymongo.ASCENDING)], unique = True)
        #mdb.votes.insert_many(votes)
        for v in votes:
            mdb.votes.insert(v)

    if scrutins:
        mdb.scrutins.create_index([('scrutin_id', pymongo.ASCENDING)], unique = True)
        mdb.scrutins.insert_many(scrutins)


    updateScrutinsStats()
    updateScrutinsSignataires()

def updateDeputesRanks():
    ops = []
    ranks = {}
    out = []
    for dep in mdb.deputes.find({'depute_actif':True}):
        compte = 'positions' in dep['stats'].keys()
        for stat in ['nbitvs','nbmots']:
            ranks[stat] = ranks.get(stat,[]) + [ (dep['depute_uid'],(dep['stats'][stat],dep['stats'][stat]) if compte else (None,None)) ]
        # stats elections
        for stat in ['inscrits','exprimes']:
            ranks['pct'+stat] = ranks.get('pct'+stat,[]) + [ (dep['depute_uid'],((2-dep['depute_election']['tour'])*100+dep['stats']['election'][stat],dep['stats']['election'][stat])) ]
        
        # stats commissions
        ranks['pctcommissions'] = ranks.get('pctcommissions',[]) + [ (dep['depute_uid'],(dep['stats']['commissions']['present'],dep['stats']['commissions']['present']) if 'commissions' in dep['stats'].keys() else (None,None)) ]
        # stats amendements
        ranks['nbamendements'] = ranks.get('nbamendements',[]) + [ (dep['depute_uid'],(dep['stats']['amendements']['rediges'],dep['stats']['amendements']['rediges']) if 'amendements' in dep['stats'].keys() else (None,None)) ]
        ranks['pctamendements'] = ranks.get('pctamendements',[]) + [ (dep['depute_uid'],(dep['stats']['amendements']['adoptes'],dep['stats']['amendements']['adoptes']) if 'amendements' in dep['stats'].keys() and 'adoptes' in dep['stats']['amendements'].keys() else (None,None)) ]
        if compte:
            for stat,val in dep['stats']['positions'].iteritems():
                ranks[stat] = ranks.get(stat,[]) + [ (dep['depute_uid'],(val,val))]
        if 'compat' in dep['stats'].keys():
            for stat,val in dep['stats']['compat'].iteritems():
                ranks['compat'+stat] = ranks.get('compat'+stat,[]) + [ (dep['depute_uid'],(val,val))]
    topflop = {}
    for rank in ranks.keys():
        topflop[rank] = {'down':{},'up':{}}
        topflop[rank]['down'] = sorted(ranks[rank],key=lambda x:x[1][0] if not x[1][0] in ('N/A',None) else -1, reverse=True)
        topflop[rank]['up'] = sorted(ranks[rank],key=lambda x:x[1][0] if not x[1][0] in ('N/A',None) else 'ZZZZ')
        topflop[rank]['down'] = dict([ (r[0],i+1) for i,r in enumerate(topflop[rank]['down']) ])
        topflop[rank]['up'] = dict([ (r[0],i+1) for i,r in enumerate(topflop[rank]['up']) ])
        
    for dep in mdb.deputes.find({'depute_actif':True}):
        dep_ranks = {'down': dict((stat,topflop[stat]['down'].get(dep['depute_uid'],None)) for stat in topflop.keys()),
                     'up': dict((stat,topflop[stat]['up'].get(dep['depute_uid'],None)) for stat in topflop.keys())}
        
        ops.append(UpdateOne({'depute_uid':dep['depute_uid']},{'$set':{'stats.ranks':dep_ranks,'stats.nonclasse':not 'positions' in dep['stats'].keys()}}))
    if ops:
        mdb.deputes.bulk_write(ops)        
    return json.dumps(ranks)


def updateAllStats():
    updateDeputesStats()
    updateGroupesStats()
    updateDeputesRanks()

def updateScrutinsSignataires():
    import re
    groupe_depute = dict((d['depute_id'],d['groupe_abrev']) for d in mdb.deputes.find({},{'depute_id':1,'groupe_abrev':1}))
    for s in mdb.scrutins.find({'scrutin_num':240}):
        desc = s['scrutin_desc'].replace('. [','')
        
        siggp = None
        sigs = None
        if 'scrutin_ref' in s.keys() and not isinstance(s['scrutin_ref'],unicode) and 'signataires' in s['scrutin_ref'].keys():
            sigs = [ normalize(sig) for sig in s['scrutin_ref']['signataires'].replace(' et ',', ').split(', ') ]
            import re
            siggp = sigs[0]
            
        if 'M. ' in s['scrutin_desc'] or 'Mme ' in s['scrutin_desc']:
            
            mtch1 = re.search(r' (M.|Mme) (.*)(,| et les amende| et l.amende)',s['scrutin_desc'],re.UNICODE)
            mtch2 =  re.search(r' (M.|Mme) (.*)( . l.article| apr.s l.article)',s['scrutin_desc'],re.UNICODE)
            

            
            siggp1 = normalize(mtch1.groups()[0]+mtch1.groups()[1]) if mtch1 else ""
            siggp2 = normalize(mtch2.groups()[0]+mtch2.groups()[1]) if mtch2 else ""
                
            siggp = siggp1 if len(siggp1)>0 and len(siggp1)<len(siggp2) else siggp2
        if siggp:
            gp = groupe_depute.get(siggp,None)
            if not gp:
                for k in groupe_depute.keys():
                    if k[:2]=='m.':
                        if k[:2]==siggp[:2] and siggp[2:]==k[len(k)-len(siggp)+2:]:
                            gp = groupe_depute[k]
                            break
                    if k[:3]=='mme':
                        if k[:3]==siggp[:3] and siggp[3:]==k[len(k)-len(siggp)+3:]:
                            gp = groupe_depute[k]
                            break
            siggp = gp
        if s['scrutin_typedetail']=='autre':
            if 'sous-amendement' in s['scrutin_desc']:
                s['scrutin_typedetail'] = 'amendement'
        
        siggp = s.get('scrutin_groupe',siggp)
        mdb.scrutins.update({'scrutin_id':s['scrutin_id']},{'$set':{'scrutin_typedetail':s['scrutin_typedetail'],'scrutin_desc':desc,'scrutin_signataires':sigs, 'scrutin_groupe':siggp}})
        mdb.votes.update_many({'scrutin_id':s['scrutin_id']},{'$set':{'scrutin_typedetail':s['scrutin_typedetail'], 'scrutin_groupe':siggp}})
        
def updateDeputesStats():
    groupes_abrev = mdb.groupes.distinct('groupe_abrev')
    groupe_depute = dict((d['depute_uid'],d['groupe_abrev']) for d in mdb.deputes.find({},{'depute_uid':1,'groupe_abrev':1}))
    pgroup = dict((g,{'$sum':'$vote_compat.'+g}) for g in mdb.groupes.distinct('groupe_abrev'))
    pgroup['_id'] = {'depute':'$depute_uid'}

    pipeline = [
        {"$group": pgroup },
    ]
   
    deputes = dict((uid,{'depute_compat':{},'depute_compat_globale':{},'depute_positions':{}}) for uid in mdb.deputes.distinct('depute_uid'))

    for compat in list(mdb.votes.aggregate(pipeline)):
        depuid = compat['_id']['depute']
        del compat['_id']
        deputes[depuid].update({'depute_compat_globale':compat})

    # comptabilisation compats
    gp_pos = dict((g,{'pour':[],'contre':[],'abstention':[],'absent':[]}) for g in groupes_abrev+['assemblee'])
    for s in mdb.scrutins.find():
        for g,pos in s['scrutin_compats'].iteritems():
            pos = pos if pos else 'absent'
            gp_pos[g][pos].append(s['scrutin_id'])
    
    for gid in ['FI','GDR','LC','LR','MODEM','NG','REM']:
        pgroup['n'] = {'$sum':1}
        pgroup['_id']['position'] ='$vote_position'
        # V1 : pipeline = [{'$match':{'$and':[{'scrutin_typedetail':'amendement'},{'vote_position':{'$ne':'absent'}}]}}] + pipeline
        pipeline2 = [{'$match':{'$and':[{'scrutin_id':{'$in':gp_pos[gid]['pour']}},{'vote_position':{'$ne':'absent'}}]}}] + pipeline
        
        for compat in list(mdb.votes.aggregate(pipeline2)):
            depuid = compat['_id']['depute']
            pos = compat['_id']['position']
            if not gid in deputes[depuid]['depute_compat'].keys():
                deputes[depuid]['depute_compat'][gid] = {}
            deputes[depuid]['depute_compat'][gid][pos] = compat[gid]
            deputes[depuid]['depute_compat'][gid]['total'] = deputes[depuid]['depute_compat'][gid].get('total',0)+compat['n']
        
    
    pipeline = [
        {"$group":{'_id': {'depute':'$depute_uid','position':'$vote_position'}, 'n':{'$sum':1}}},
        {"$sort": SON([("_id.depute",1)])}
    ]
    
    for p in list(mdb.votes.aggregate(pipeline)):
        uid = p['_id']['depute']
        deputes[uid]['depute_positions'][p['_id']['position']] = p['n']
    ops = []
    
    
    gp_nbvotes = {}
    for g in groupes_abrev:
        pipeline = [ {'$redact': {'$cond': [{'$lt': ['$scrutin_positions.'+g+'.absent', '$scrutin_positions.'+g+'.total']}, '$$KEEP', '$$PRUNE']}}]
        gp_nbvotes[g] = len(list(mdb.scrutins.aggregate(pipeline)))
    
    for uid,dep in deputes.iteritems():
        dep['depute_positions']['total'] = dep['depute_positions'].get('pour',0)+dep['depute_positions'].get('contre',0)+dep['depute_positions'].get('abstention',0)+dep['depute_positions'].get('absent',0)
        dep['depute_positions']['exprimes'] = dep['depute_positions']['total'] - dep['depute_positions'].get('absent',0)
        dep['depute_positions']['dissidence'] = dep['depute_positions']['exprimes'] - dep['depute_compat_globale'].get(groupe_depute[uid],0)

        # stats positions
        dep['stats.positions'] = dict((p,0) for p in ('pour','contre','exprimes','absent','abstention','dissidence'))
        for pos in dep['stats.positions'].keys():
            if dep['depute_positions']['exprimes']>0:
                num = dep['depute_positions']['exprimes'] if not pos in ('exprimes','absent') else dep['depute_positions']['total']
                dep['stats.positions'][pos] = round(100*float(dep['depute_positions'].get(pos,0)) / num,3)
        # stats compat
        dep['stats.compat'] = dict((g,0) for g in groupes_abrev)
        for g,v in dep['depute_compat'].iteritems():
            #if dep['depute_positions']['exprimes']>0:
                #dep['stats.compat'][g] = round(100*float(v) / gp_nbvotes[g],3)
            if dep['stats.positions']['exprimes']>seuil_compat:
                dep['stats.compat'][g] = round(100*float(v.get('pour',0)) / v['total'],3)
            else:
                dep['stats.compat'][g] = None
        dep['stats.compat_sort'] = [ dict(g=g,p=p) for g,p in sorted(dep['stats.compat'].items(), key=lambda x:x[1], reverse=True) ]
    
    # Vote pour / contre amendements autres groupes
    pgroup = {}
    pgroup['n'] = {'$sum':1}
    pgroup['_id'] = { 'uid':'$depute_uid','grouped':'$groupe_abrev','groupes':'$scrutin_groupe','position':'$vote_position'}
    redact = {
        "$redact": {
            "$cond": [
                { "$ne": [ "$scrutin_groupe", "$groupe_abrev" ] },
                "$$KEEP",
                "$$PRUNE"
            ]
        }
    }
    pipeline = [{'$match':{'scrutin_typedetail':'amendement'}},{'$group':pgroup}]
    voteamtdeps = {}
    
    for p in mdb.votes.aggregate(pipeline):
        if p['_id']['position']!='absent':
            uid = p['_id']['uid']
            pos = p['_id']['position']
            if p['_id']['groupes']!=None and p['_id']['grouped']!=p['_id']['groupes']:
                if not uid in voteamtdeps:
                    voteamtdeps[uid] = {'pour':0,'contre':0,'abstention':0,'total':0}
                voteamtdeps[uid][pos] += p['n']
                voteamtdeps[uid]['total'] += p['n']
    count = 0                
    for uid,data in voteamtdeps.iteritems():
        if data['total']>0:
            for pos in ('pour','contre','abstention'):
                voteamtdeps[uid]['pct'+pos] = round(100*float(data[pos])/data['total'],2)

    # votes scrutins clés
    votes_cles = {}
    for v in mdb.votes.find({'scrutin_num':{'$in':scrutins_cles.keys()}},{'scrutin_num':1,'depute_uid':1,'vote_position':1}):
        if not v['depute_uid'] in votes_cles.keys():
            votes_cles[v['depute_uid']] = {}
        votes_cles[v['depute_uid']][str(v['scrutin_num'])] = v['vote_position']

    # ASSEMBLAGE FINAL
    for d in mdb.deputes.find():
        uid = d['depute_uid']
        if uid in deputes.keys():
            dep = deputes[uid]
        else:
            dep = {}
        # stats election
        dep['stats.election'] = {'exprimes':round(100*float(d['depute_election']['voix'])/d['depute_election']['exprimes'],2),
                                 'inscrits':round(100*float(d['depute_election']['voix'])/d['depute_election']['inscrits'],2)}

        # positions votes cles
        if uid in votes_cles.keys():
            dep['depute_votes_cles'] = votes_cles[uid]
        # stats votes amds autres groupes
        if uid in voteamtdeps.keys():
            dep['stats.votesamdements'] = voteamtdeps[uid]
        # stats commissions
        statsc = {'present':0,'absent':0,'excuse':0,'total':0}
        if 'depute_presences_commissions' in d.keys():
            for c in d['depute_presences_commissions'].values():
                for pos in ('present','absent','excuse'):
                    statsc[pos] += c[pos]
                    statsc['total'] += c[pos]

            if statsc['total']>0:
                dep['stats.commissions'] = {}
                for pos in ('present','absent','excuse'):
                    dep['stats.commissions'][pos] = round(100*float(statsc[pos])/statsc['total'],2)
        # stats amendements
        if 'depute_amendements' in d.keys():
            dep['stats.amendements'] = {'rediges':d['depute_amendements']['rediges']}
            if d['depute_amendements']['rediges']>0:
                dep['stats.amendements']['adoptes']=round(100*float(d['depute_amendements']['adoptes'])/d['depute_amendements']['rediges'],2)
        ops.append(UpdateOne({'depute_uid':uid},{'$set':dep}))
        
    
    if ops:
        mdb.deputes.bulk_write(ops)
        mdb.deputes.update_many({'groupe_abrev':'NI'},{'$set':{'stats.positions.dissidence':'N/A'}})

    return json.dumps(gp_nbvotes)


def updateScrutinsStats():
    if 'rebuild' in request.args:
        mdb.votes.update_many({},{'$set':{'vote_compat':None}})

        
    
    # decomptes positions par scrutin, par groupe, par position
    pipeline = [
        {"$match":{'vote_compat':None}},
        {"$group":{'_id': {'position':'$vote_position','scrutin':'$scrutin_id','groupe':'$groupe_abrev'},'total':{'$sum':1}}},
        {"$sort": SON([("_id.scrutin",1),("_id.groupe",1),("total",-1),("_id.position",1)])},

    ]

    # Stats scrutins
    scrutins = {}

    scrutin_id = None
    for r in list(mdb.votes.aggregate(pipeline)):
        if r['_id']['scrutin'] != scrutin_id:
            scrutin_id = r['_id']['scrutin']
            scrutins[scrutin_id] = {}
            scrutins[scrutin_id]['positions'] = {'assemblee':{'pour':0,'contre':0,'abstention':0,'absent':0}}
            scrutins[scrutin_id]['compats'] = {}
            groupe = None
        position = r['_id']['position']
        if r['_id']['groupe'] != groupe:
            groupe = r['_id']['groupe']
            scrutins[scrutin_id]['positions'][groupe] = {'position':''}
            scrutins[scrutin_id]['compats'][groupe] = ''
        if position != 'absent' and scrutins[scrutin_id]['positions'][groupe]['position']=='':
            scrutins[scrutin_id]['positions'][groupe]['position'] = position
            scrutins[scrutin_id]['compats'][groupe] = position
        scrutins[scrutin_id]['positions'][groupe][position] = r['total']
        scrutins[scrutin_id]['positions']['assemblee'][position] += r['total']
    # Complements stats
    groupes = mdb.groupes.distinct('groupe_abrev')
    groupe_init = dict((abr,0) for abr in groupes+['assemblee'])
    ops = []
    for sid,scrutin in scrutins.iteritems():
        tri = sorted([(k,v) for k,v in scrutin['positions']['assemblee'].iteritems() if k!='absent'], key=lambda x:x[1], reverse=True)
        asspos = tri[0][0] if len(tri)>0 else ''
        scrutin['positions']['assemblee']['position'] = asspos
        scrutin['compats']['assemblee'] = asspos

        compatpos = { 'absent':groupe_init }
        for pos in ('pour','contre','abstention'):
            compatpos[pos] = dict((g,1 if scrutin['compats'][g]==pos else 0) for g in groupes+['assemblee'])
        scrutin['compatpos'] = compatpos
        for g in groupes+['assemblee']:
            scrutin['positions'][g]['exprimes'] = scrutin['positions'][g].get('pour',0) + scrutin['positions'][g].get('contre',0) + scrutin['positions'][g].get('abstention',0)
            scrutin['positions'][g]['total'] = scrutin['positions'][g]['exprimes'] + scrutin['positions'][g].get('absent',0)
        scrutin['sort'] = u'adopté' if scrutin['positions']['assemblee']['position']=='pour' else u'rejeté'
        ops.append(UpdateOne({'scrutin_id':sid},{'$set':{'scrutin_positions':scrutin['positions'], 'scrutin_compats':scrutin['compats'], 'scrutin_sort':scrutin['sort']}}))

    if ops:
        mdb.scrutins.bulk_write(ops)
    ops = []
    for v in mdb.votes.find({'scrutin_id':{'$in':scrutins.keys()}}):
        ops.append(UpdateOne({'vote_id':v['vote_id']},{'$set':{'vote_compat':scrutins[v['scrutin_id']]['compatpos'][v['vote_position']]}}))
    if ops:
        mdb.votes.bulk_write(ops)
        


    return "OK"

def updateGroupesStats():
    #pipeline = [
    #    {"$group":{'_id': {'groupe':'$groupe_abrev','csp':'$depute_csp'}, 'n':{'$sum':1}}},
    #    {"$sort": SON([("n",-1)])}
    #]
    #return json.dumps(list(mdb.deputes.aggregate(pipeline)))
    
    pgroup = dict((g,{'$sum':'$vote_compat.'+g}) for g in mdb.groupes.distinct('groupe_abrev'))
    pgroup['_id'] = {'groupe':'$groupe_abrev'}
    pipeline = [
        {"$group": pgroup },
    ]
    groupes = {}
    for compat in list(mdb.votes.aggregate(pipeline)):
        groupe = compat['_id']['groupe']
        del compat['_id']
        groupes[groupe] = {'groupe_compat_globale':compat, 'groupe_positions':{} }        

    pgroup = dict((g+'_pour',{'$sum':'$depute_compat.'+g+'.pour'}) for g in ['FI','GDR','LC','LR','MODEM','NG','REM'])
    pgroup.update(dict((g+'_total',{'$sum':'$depute_compat.'+g+'.total'}) for g in ['FI','GDR','LC','LR','MODEM','NG','REM']))
    pgroup['_id'] = {'groupe':'$groupe_abrev'}
    pipeline = [
        {"$group": pgroup },
    ]
    for cpt in mdb.deputes.aggregate(pipeline):
        gp = cpt['_id']['groupe']
        _comp = {}
        for g in ['FI','GDR','LC','LR','MODEM','NG','REM']:
            _comp[g] = {'pour':cpt[g+'_pour'],'total':cpt[g+'_total']}
        groupes[gp]['groupe_compat'] = _comp


    pipeline = [
        {"$group":{'_id': {'groupe':'$groupe_abrev','position':'$vote_position'}, 'n':{'$sum':1}}},
        {"$sort": SON([("_id.groupe",1)])}
    ]

    for p in list(mdb.votes.aggregate(pipeline)):
        gid = p['_id']['groupe']
        groupes[gid]['groupe_positions'][p['_id']['position']] = p['n']
    ops = []
    
    for gid,g in groupes.iteritems():
        g['groupe_positions']['total'] = sum(g['groupe_positions'].values())
        g['groupe_positions']['exprimes'] = g['groupe_positions']['total'] - g['groupe_positions'].get('absent',0)
        g['groupe_positions']['dissidence'] = g['groupe_positions']['exprimes'] - g['groupe_compat_globale'].get(gid,0)

        # stats positions
        g['stats.positions'] = dict((p,0) for p in ('pour','contre','exprimes','absent','abstention','dissidence'))
        for pos in g['stats.positions'].keys():
            if g['groupe_positions']['exprimes']>0:
                g['stats.positions'][pos] = round(100*float(g['groupe_positions'].get(pos,0)) / g['groupe_positions']['total'],3)
        # stats compat
        g['stats.compat'] = dict((_g,0) for _g in g['groupe_compat'].keys())
        for _g,_v in g['groupe_compat'].iteritems():
            if g['groupe_positions']['exprimes']>0:
                g['stats.compat'][_g] = round(100*float(_v['pour']) / _v['total'],3)
        g['stats.compat_sort'] = [ dict(g=_g,p=_p) for _g,_p in sorted(g['stats.compat'].items(), key=lambda x:x[1], reverse=True) ]
        
        
        stats = {'classeage':{},'csp':{}}
        for d in mdb.deputes.find({'groupe_abrev':gid},{'depute_id':1,'depute_csp':1,'depute_classeage':1}):
            sexe = 'F' if d['depute_id'][0:3]=='mme' else 'H'
            stats['csp'][d['depute_csp']] = stats['csp'].get(d['depute_csp'],0) + 1
            if not d['depute_classeage'] in stats['classeage'].keys():
                stats['classeage'][d['depute_classeage']]={'H':0,'F':0}
            stats['classeage'][d['depute_classeage']][sexe] += 1
        g.update({'stats.csp':sorted(stats['csp'].items(),key=lambda x:x[1],reverse=True),
                  'stats.classeage':sorted(stats['classeage'].items(),key=lambda x:x[0],reverse=True)
                 })
        ops.append(UpdateOne({'groupe_abrev':gid},{'$set':g}))

    if ops:
        mdb.groupes.bulk_write(ops)

    return json.dumps(groupes)

def updateScrutinsCles():
    from openpyxl import load_workbook
    from cStringIO import StringIO
    import requests
    
    r =requests.get("https://docs.google.com/spreadsheets/d/1lZ5aMIaglRh6_BK66AYkXzi9guJaMij9RlpSoO4dnIw/export?format=xlsx&id=1lZ5aMIaglRh6_BK66AYkXzi9guJaMij9RlpSoO4dnIw")
    f = StringIO(r.content)
    wb = load_workbook(f)
    for niv in range(2):
        ws = wb[wb.sheetnames[niv]]
        elts = []
        for j,row in enumerate(ws.iter_rows(min_row=2)):
            elts.append([a.value for a in list(row[:4])]+[niv])
    
    for e in elts:
        if e[0] != None:
            mdb.scrutinscles.update({'num':int(e[0])},{'$set':dict(num=int(e[0]),nom=e[1],desc=e[2],theme=e[3],niveau=e[4])},upsert=True)
    return "ok"
