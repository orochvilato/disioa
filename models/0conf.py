# -*- coding: utf-8 -*-

colors = { 'FI':'rgba(12,195,214,1)',
          'NI':'rgba(191,191,191,1)',
          'NG':'rgba(234,52,208,1)',
          'REM':'rgba(255,192,0,1)',
          'MODEM':'rgba(237,125,49,1)',
          'LR':'rgba(47,85,151,1)',
          'LC':'rgba(112,48,160,1)',
          'GDR':'rgba(192,0,0,1)'
        }

departements_ids = {u'Haute-Sa\xf4ne': '070', u'Tarn': '081', u'Paris': '075', u'Fran\xe7ais \xe9tablis hors de France': '999', u'Corse-du-Sud': '02A', u'Ard\xe8che': '007', u'Doubs': '025', u'Calvados': '014', u'Var': '083', u'Haut-Rhin': '068', u'Meurthe-et-Moselle': '054', u'Haute-Marne': '052', u'Moselle': '057', u'Loz\xe8re': '048', u'Gironde': '033', u'Haute-Loire': '043', u"Val-d'Oise": '095', u'Eure': '027', u'Jura': '039', u'Loir-et-Cher': '041', u'Aveyron': '012', u'Ain': '001', u'Seine-et-Marne': '077', u'Yonne': '089', u'Mayenne': '053', u'Val-de-Marne': '094', u'Alpes-Maritimes': '006', u'Corr\xe8ze': '019', u'Gard': '030', u"C\xf4tes-d'Armor": '022', u'Haute-Vienne': '087', 'Saint-Pierre-et-Miquelon': '975', u'Saint-Barth\xe9lemy et Saint-Martin': '977', u'Ari\xe8ge': '009', u'Vend\xe9e': '085', u'Alpes-de-Haute-Provence': '004', u'Sarthe': '072', u'Loire-Atlantique': '044', u"C\xf4te-d'Or": '021', u'Creuse': '023', u'Oise': '060', u'Finist\xe8re': '029', u'Lot-et-Garonne': '047', u'Cantal': '015', u'Vienne': '086', u'Yvelines': '078', u'Bouches-du-Rh\xf4ne': '013', u'Meuse': '055', u'Orne': '061', u'Indre-et-Loire': '037', u'Maine-et-Loire': '049', u'Vaucluse': '084', u'Hauts-de-Seine': '092', u'Deux-S\xe8vres': '079', u'Loire': '042', u'Tarn-et-Garonne': '082', u'Essonne': '091', u'Indre': '036', u'Pas-de-Calais': '062', u'Aude': '011', u'Hautes-Alpes': '005', u'Eure-et-Loir': '028', u'Morbihan': '056', u'Nouvelle-Cal\xe9donie': '988', u'Lot': '046', u'Haute-Corse': '02B', u'Charente': '016', u'Savoie': '073', u'Loiret': '045', u'Manche': '050', u'Pyr\xe9n\xe9es-Atlantiques': '064', u'Puy-de-D\xf4me': '063', u'Ille-et-Vilaine': '035', u'Allier': '003', u'Vosges': '088', u'Marne': '051', u'Rh\xf4ne': '069', u'Sa\xf4ne-et-Loire': '071', u'Pyr\xe9n\xe9es-Orientales': '066', u'Aube': '010', u'Is\xe8re': '038', u'Aisne': '002', u'Haute-Savoie': '074', u'Haute-Garonne': '031', u'Gers': '032', u'Nord': '059', u'Charente-Maritime': '017', u'R\xe9union': '974', u'Landes': '040', u'Hautes-Pyr\xe9n\xe9es': '065', u'H\xe9rault': '034', u'Dr\xf4me': '026', u'Martinique': '972', u'Wallis-et-Futuna': '986', u'Somme': '080', u'Ni\xe8vre': '058', u'Seine-Saint-Denis': '093', u'Mayotte': '976', u'Seine-Maritime': '076', u'Dordogne': '024', u'Guyane': '973', u'Guadeloupe': '971', u'Ardennes': '008', u'Cher': '018', u'Polyn\xe9sie Fran\xe7aise': '987', u'Bas-Rhin': '067', u'Territoire de Belfort': '090'}



nuages_excl =('ministre','monsieur',u'collègue','madame','premier','cher','effet','commission','parlementaire',u'député','deux','amendement','sujet','texte','an','fois','lieu','trois','mois','rapporteur','cinq','quant','quel','dix','nom','quelle',u'deuxième','hui','vingt','point','question','y','ne','pas','dire','plus','fait','bien','dit',u'président')

def getScrutinsCles():
    from openpyxl import load_workbook
    from cStringIO import StringIO
    import requests
    
    r =requests.get("https://docs.google.com/spreadsheets/d/1lZ5aMIaglRh6_BK66AYkXzi9guJaMij9RlpSoO4dnIw/export?format=xlsx&id=1lZ5aMIaglRh6_BK66AYkXzi9guJaMij9RlpSoO4dnIw")
    f = StringIO(r.content)
    wb = load_workbook(f)
    ws = wb['scrutins']
    elts = []
    for j,row in enumerate(ws.iter_rows(min_row=2)):
        elts.append([a.value for a in row])
    scles = {}
    for e in elts:
        if e[0] != None:
            scles[int(e[0])] = dict(num=int(e[0]),nom=e[1],desc=e[2],theme=e[3],lien=e[4])
    return scles

scrutins_cles = cache.disk('scrutins_cles',lambda:getScrutinsCles(),time_expire=3600*24)

csp_liste = [(u"Cadres et professions intellectuelles sup\u00e9rieures",u"Cadres, Prof. Sup."), (u"Artisans, commer\u00e7ants et chefs d'entreprise",u"Artisants, Chefs d'entrep."), (u"Agriculteurs exploitants",u"Agriculteurs exploitants"),(u"Professions Interm\u00e9diaires",u"Professions Interm\u00e9diaires"),(u"Employ\u00e9s",u"Employ\u00e9s"),(u"Ouvriers",u"Ouvriers"),(u"Retrait\u00e9s",u"Retrait\u00e9s"),(u"Autres (y compris inconnu et sans profession d\u00e9clar\u00e9e)",u"Autres")]

classeage_liste = ["70-80 ans", "60-70 ans", "50-60 ans", "40-50 ans", "30-40 ans", "20-30 ans"]

seuil_compat = 20
